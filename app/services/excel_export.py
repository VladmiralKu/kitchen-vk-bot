from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time, timezone
from pathlib import Path
from statistics import median
from zoneinfo import ZoneInfo

from openpyxl import Workbook
from openpyxl.styles import Font


EXPORT_DIR = Path("/tmp/kitchen-vk-bot-exports")


@dataclass(frozen=True)
class ExportPeriod:
    start_utc: datetime
    end_utc: datetime
    label: str


def parse_export_period(args: list[str], timezone_name: str) -> ExportPeriod:
    tz = ZoneInfo(timezone_name)
    today = datetime.now(tz).date()

    if not args or args == ["today"]:
        start_date = today
        end_date = today
        label = "today"
    elif args == ["yesterday"]:
        start_date = today.fromordinal(today.toordinal() - 1)
        end_date = start_date
        label = "yesterday"
    elif len(args) == 2:
        start_date = date.fromisoformat(args[0])
        end_date = date.fromisoformat(args[1])
        label = f"{start_date}_{end_date}"
    else:
        raise ValueError("Используйте /export today или /export YYYY-MM-DD YYYY-MM-DD")

    start_local = datetime.combine(start_date, time.min, tzinfo=tz)
    end_local = datetime.combine(end_date, time.max, tzinfo=tz)
    return ExportPeriod(start_local.astimezone(timezone.utc), end_local.astimezone(timezone.utc), label)


def build_export_workbook(data: dict, timezone_name: str) -> Workbook:
    wb = Workbook()
    default_sheet = wb.active
    wb.remove(default_sheet)

    _write_sheet(
        wb,
        "Orders",
        [
            "order_no",
            "status",
            "table_number",
            "waiter_name",
            "raw_text",
            "comment",
            "created_at",
            "ready_at",
            "completed_at",
            "total_ready_seconds",
            "total_ready_minutes",
        ],
        data.get("orders", []),
        timezone_name,
    )
    _write_sheet(
        wb,
        "Items",
        [
            "order_no",
            "item_index",
            "quantity",
            "name",
            "status",
            "ready_by_name",
            "created_at",
            "ready_at",
            "ready_seconds",
            "ready_minutes",
        ],
        data.get("items", []),
        timezone_name,
    )
    _write_sheet(wb, "Events", ["order_no", "item_name", "user_name", "event_type", "created_at", "payload"], data.get("events", []), timezone_name)
    _write_sheet(wb, "Users", ["vk_user_id", "display_name", "role", "status", "created_at"], data.get("users", []), timezone_name)
    _write_sheet(wb, "Stats", ["date", "orders_count", "avg_ready_minutes", "max_ready_minutes", "min_ready_minutes", "median_ready_minutes"], data.get("stats", []), timezone_name)
    return wb


async def create_export_file(session, period: ExportPeriod, timezone_name: str) -> Path:
    from sqlalchemy import select
    from sqlalchemy.orm import selectinload

    from app.models.order import Order
    from app.models.order_event import OrderEvent
    from app.models.user import User

    orders_result = await session.execute(
        select(Order)
        .options(selectinload(Order.items), selectinload(Order.waiter))
        .where(Order.created_at >= period.start_utc, Order.created_at <= period.end_utc)
        .order_by(Order.created_at)
    )
    orders = list(orders_result.scalars().unique())

    users_result = await session.execute(select(User).order_by(User.created_at))
    users = list(users_result.scalars())
    users_by_id = {user.id: user for user in users}
    orders_by_id = {order.id: order for order in orders}

    events_result = await session.execute(
        select(OrderEvent)
        .where(OrderEvent.created_at >= period.start_utc, OrderEvent.created_at <= period.end_utc)
        .order_by(OrderEvent.created_at)
    )
    events = list(events_result.scalars())

    data = {
        "orders": [_order_row(order) for order in orders],
        "items": [_item_row(order, item, users_by_id) for order in orders for item in order.items],
        "events": [_event_row(event, orders_by_id, users_by_id) for event in events],
        "users": [_user_row(user) for user in users],
        "stats": [_stats_row(orders, period, timezone_name)],
    }

    workbook = build_export_workbook(data, timezone_name)
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    path = EXPORT_DIR / f"kitchen_orders_{period.label}_{datetime.now(timezone.utc).strftime('%Y%m%d%H%M%S')}.xlsx"
    workbook.save(path)
    return path


def _write_sheet(wb: Workbook, title: str, headers: list[str], rows: list[dict], timezone_name: str) -> None:
    ws = wb.create_sheet(title)
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
    for row in rows:
        ws.append([_format_value(row.get(header), timezone_name) for header in headers])
    for column in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column)
        ws.column_dimensions[column[0].column_letter].width = min(max(width + 2, 12), 60)


def _format_value(value, timezone_name: str):
    if isinstance(value, datetime):
        tz = ZoneInfo(timezone_name)
        return value.astimezone(tz).strftime("%Y-%m-%d %H:%M:%S")
    return value


def _order_row(order) -> dict:
    return {
        "order_no": order.order_no,
        "status": order.status,
        "table_number": order.table_number,
        "waiter_name": getattr(order.waiter, "display_name", None),
        "raw_text": order.raw_text,
        "comment": order.comment,
        "created_at": order.created_at,
        "ready_at": order.ready_at,
        "completed_at": order.completed_at,
        "total_ready_seconds": order.total_ready_seconds,
        "total_ready_minutes": round(order.total_ready_seconds / 60, 2) if order.total_ready_seconds is not None else None,
    }


def _item_row(order, item, users_by_id: dict) -> dict:
    ready_by = users_by_id.get(item.ready_by)
    return {
        "order_no": order.order_no,
        "item_index": item.position_index,
        "quantity": float(item.quantity),
        "name": item.name,
        "status": item.status,
        "ready_by_name": getattr(ready_by, "display_name", None),
        "created_at": item.created_at,
        "ready_at": item.ready_at,
        "ready_seconds": item.ready_seconds,
        "ready_minutes": round(item.ready_seconds / 60, 2) if item.ready_seconds is not None else None,
    }


def _event_row(event, orders_by_id: dict, users_by_id: dict) -> dict:
    order = orders_by_id.get(event.order_id)
    user = users_by_id.get(event.user_id)
    return {
        "order_no": getattr(order, "order_no", None),
        "item_name": None,
        "user_name": getattr(user, "display_name", None),
        "event_type": event.event_type,
        "created_at": event.created_at,
        "payload": str(event.payload or {}),
    }


def _user_row(user) -> dict:
    return {
        "vk_user_id": user.vk_user_id,
        "display_name": user.display_name,
        "role": user.role,
        "status": user.status,
        "created_at": user.created_at,
    }


def _stats_row(orders: list, period: ExportPeriod, timezone_name: str) -> dict:
    ready_minutes = [order.total_ready_seconds / 60 for order in orders if order.total_ready_seconds is not None]
    period_date = period.start_utc.astimezone(ZoneInfo(timezone_name)).date().isoformat()
    return {
        "date": period_date,
        "orders_count": len(orders),
        "avg_ready_minutes": round(sum(ready_minutes) / len(ready_minutes), 2) if ready_minutes else None,
        "max_ready_minutes": round(max(ready_minutes), 2) if ready_minutes else None,
        "min_ready_minutes": round(min(ready_minutes), 2) if ready_minutes else None,
        "median_ready_minutes": round(median(ready_minutes), 2) if ready_minutes else None,
    }
